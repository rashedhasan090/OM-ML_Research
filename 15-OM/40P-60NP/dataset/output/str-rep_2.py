import os
import openpyxl

def find_replace_excel(file_path, sheet_name, replacements):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the specified sheet
    sheet = workbook[sheet_name]

    # Iterate through each row in the sheet
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        # Iterate through each cell in the row
        for cell in row:
            if cell is not None and isinstance(cell, str):
                # Replace the words in the cell value
                for find_word, replace_word in replacements.items():
                    cell = cell.replace(find_word, replace_word)

    # Save the modified workbook
    workbook.save(file_path)
    print(f"Modifications saved in {file_path}")

# Get the current directory
current_directory = os.getcwd()

# Get a list of all .xlsx files in the current directory
xlsx_files = [file for file in os.listdir(current_directory) if file.endswith(".xlsx")]

# Iterate over each .xlsx file
for file in xlsx_files:
    # Specify the path to the Excel file
    file_path = os.path.join(current_directory, file)
    
    # Specify the sheet name and replacements dictionary
    sheet_name = "Sheet1"  # Update with your sheet name
    replacements = {
    "ONDELETECASCADEONUPDATECASCADE": " ON DELETE CASCADE ON UPDATE CASCADE ",
    "CREATETABLE": "CREATE TABLE ",
    "parentparent": "parent parent",
    "Associationsrc": " Association src",
    "onesig": " one sig ",
    "moduleOM_name:0openDeclarationonesig": "module OM_name:0 open Declaration one sig ",
    "MappingStrategyof": " Mapping Strategy of ",
    "PRIMARYKEY": " PRIMARY KEY ",
    "FOREIGNKEY": " FOREIGN KEY ",
    "NOTNULL": "NOT NULL",
    "REFERENCES": " REFERENCES ",
    "KEY": " KEY ",
    "extends": " extends ",
    "predshowrunshow": " pred show run show ",
    "ADDCONSTRAINT": " ADD CONSTRAINT ",
    "extendsClass": " extends Class ",
    "moduleOM_name:0": "module OM_name:0",
    "dst_multiplicity": " dst_multiplicity ",
    "src_multiplicity": " src_multiplicity ",
    "PRIMARYKEY": " PRIMARY KEY ",
    "MappingStrategyofTable": " Mapping Strategy of Table ",
    "extendsAssociation": "extends Association",
    "Nonoparent": "No no parent",
    "Declarationone": "Declaration one",
    "oneparent"     :  " one parent ",
    "moduleOM_name0": "moduleOM_name0",
    "ALTERTABLE"    :  " ALTER TABLE ",
    "NOT" : " NOT ",
    "AssociationStrategyfor" : " Association Strategy for ",
    "parentin" : " parent in ",
    "noparentisAbstract=No" : "no parent is Abstract = No ",
    "MappingStrategyfor" : " Mapping Strategy for ",
    "USEOM" : " USE OM ",
    "openDeclaration" : "open Declaration",
    "Class attrSet": "Class attrSet",
    "noparent" : " no parent "
    }

    # Apply find and replace in the Excel file
    find_replace_excel(file_path, sheet_name, replacements)
