import openpyxl

def process_excel_file(input_file, output_file):
    # Load the input Excel file
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Find the column index of the OM_Prediction column
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        prediction_col_idx = header_row.index('OM_Prediction') + 1
    except ValueError:
        print("OM_Prediction column not found in the Excel file.")
        return

    # Create a list to store modified rows
    modified_rows = []

    # Iterate through each row and process the OM_Prediction column
    for row in sheet.iter_rows(min_row=2, values_only=True):
        prediction = row[prediction_col_idx - 1]
        if prediction.startswith("P,"):
            modified_rows.append(row[:prediction_col_idx - 1] + (1,) + row[prediction_col_idx:])
        elif prediction.startswith("NP,"):
            modified_rows.append(row[:prediction_col_idx - 1] + (0,) + row[prediction_col_idx:])
    
    # Write modified rows back to the Excel file
    for idx, modified_row in enumerate(modified_rows, start=2):
        for col_idx, cell_value in enumerate(modified_row, start=1):
            sheet.cell(row=idx, column=col_idx, value=cell_value)

    # Save changes to a new Excel file
    wb.save(output_file)

# Usage example
input_file = 'random_set_3.xlsx'
output_file = 'random_set_3.0.xlsx'

process_excel_file(input_file, output_file)
