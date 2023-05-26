import openpyxl

def replace_strings_in_excel(input_file, output_file, replacements):
    # Load the input Excel file
    wb = openpyxl.load_workbook(input_file)

    # Iterate over all sheets in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Iterate over all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if the cell value matches any of the replacements
                    for search_string, replacement_string in replacements.items():
                        if str(cell.value) == search_string:
                            # Replace the value with the given string
                            cell.value = replacement_string

    # Save the modified workbook to the output file
    wb.save(output_file)
    print("Replacement completed. Output file saved as:", output_file)


# Example usage
input_file = "12_OM_60.xlsx"
output_file = "12_OM_60_label.xlsx"
replacements = {
    r"ONDELETECASCADEONUPDATECASCADE": " ON DELETE CASCADE ON UPDATE CASCADE ",
    r"CREATETABLE": "CREATE TABLE ",
    r"parentparent": "parent parent",
    r"Associationsrc": " Association src",
    r"onesig" : " one sig ",
    r"moduleOM_name:0openDeclarationonesig" : "module OM_name:0 open Declaration one sig ",
    r"MappingStrategyof":" Mapping Strategy of ",
    r"PRIMARYKEY" : " PRIMARY KEY ",
    r"FOREIGNKEY": " FOREIGN KEY ",
    r"NOTNULL": "NOT NULL",
    r"REFERENCES": " REFERENCES ",
    r"KEY" : " KEY ",
    r"extends" : " extends ",
    r"predshowrunshow" : " pred show run show ",
    r"ADDCONSTRAINT"  : " ADD CONSTRAINT ",
    r"extendsClass" : " extends Class ",
    r"moduleOM_name:0" : "module OM_name:0",
    r"dst_multiplicity": " dst_multiplicity ",
    r"src_multiplicity": " src_multiplicity ",
    r"PRIMARYKEY" : " PRIMARY KEY ",
    r"MappingStrategyofTable": " Mapping Strategy of Table "

}

replace_strings_in_excel(input_file, output_file, replacements)
