import openpyxl

def find_replace_excel(file_path, sheet_name, replacements, output_file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the specified sheet
    sheet = workbook[sheet_name]

    # Create a new workbook for output
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    # Iterate through each row in the sheet
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        # Create a new row for output
        output_row = []

        # Iterate through each cell in the row
        for cell in row:
            if cell is not None and isinstance(cell, str):
                # Replace the words in the cell value
                for find_word, replace_word in replacements.items():
                    cell = cell.replace(find_word, replace_word)

            # Add the modified cell value to the output row
            output_row.append(cell)

        # Append the output row to the output sheet
        output_sheet.append(output_row)

    # Save the modified workbook to a new file
    output_workbook.save(output_file_path)
    print(f"Modified file saved as {output_file_path}")

# Example usage
file_path = "10_OM_50.xlsx"
sheet_name = "Sheet1"
replacements = {
    "ONDELETECASCADEONUPDATECASCADE": " ON DELETE CASCADE ON UPDATE CASCADE ",
    "CREATETABLE": "CREATE TABLE ",
    "parentparent": "parent parent",
    "Associationsrc": " Association src",
    "onesig": " one sig ",
    "moduleOM_name:0openDeclarationonesig": "module OM_name:0 open Declaration one sig ",
    "MappingStrategyof": " Mapping Strategy of ",
    "PRIMARYKEY": " PRIMARY KEY ",
    "FOREIGNKEY": " FOREIGN KEY ",
    "NOTNULL": "NOT NULL",
    "REFERENCES": " REFERENCES ",
    "KEY": " KEY ",
    "extends": " extends ",
    "predshowrunshow": " pred show run show ",
    "ADDCONSTRAINT": " ADD CONSTRAINT ",
    "extendsClass": " extends Class ",
    "moduleOM_name:0": "module OM_name:0",
    "dst_multiplicity": " dst_multiplicity ",
    "src_multiplicity": " src_multiplicity ",
    "PRIMARYKEY": " PRIMARY KEY ",
    "MappingStrategyofTable": " Mapping Strategy of Table ",
    "extendsAssociation": "extends Association",
    "Nonoparent": "No no parent",
    "Declarationone": "Declaration one",
    "oneparent"     :  " one parent ",
    "moduleOM_name0": "moduleOM_name0",
    "ALTERTABLE"    :  " ALTER TABLE ",
    "NOT" : " NOT ",
    "AssociationStrategyfor" : " Association Strategy for ",
    "parentin" : " parent in ",
    "noparentisAbstract=No" : "no parent is Abstract = No ",
    "MappingStrategyfor" : " Mapping Strategy for ",
    "USEOM" : " USE OM ",
    "openDeclaration" : "open Declaration",
    "Class attrSet": "Class attrSet"
}

output_file_path = "10_OM_50_space.xlsx"
find_replace_excel(file_path, sheet_name, replacements, output_file_path)
