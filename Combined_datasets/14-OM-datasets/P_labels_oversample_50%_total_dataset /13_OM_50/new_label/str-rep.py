import openpyxl

def find_replace_excel(file_path, sheet_name, find_words, replace_words, output_file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the specified sheet
    sheet = workbook[sheet_name]

    # Create a new workbook for output
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    # Iterate through each row in the sheet
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        # Create a new row for output
        output_row = []

        # Iterate through each cell in the row
        for cell in row:
            if cell is not None and isinstance(cell, str):
                # Replace the find words with replace words in the cell value
                for find_word, replace_word in zip(find_words, replace_words):
                    cell = cell.replace(find_word, replace_word)

            # Add the modified cell value to the output row
            output_row.append(cell)

        # Append the output row to the output sheet
        output_sheet.append(output_row)

    # Save the output workbook
    output_workbook.save(output_file_path)



# Example usage
file_path = "13_OM_50.xlsx"
sheet_name = "Sheet1"
find_words = ["ONDELETECASCADEONUPDATECASCADE", ]
replace_words = [" ON DELETE CASCADE ON UPDATE CASCADE "]
output_file_path = "13_OM_50_1.xlsx"
find_replace_excel(file_path, sheet_name, find_words, replace_words, output_file_path)
